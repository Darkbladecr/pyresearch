from datetime import datetime
from sets import Set
from collections import OrderedDict
from operator import itemgetter
from openpyxl import load_workbook, Workbook
from openpyxl.chart import LineChart, Reference, Series
from Bio import Entrez
from iso639 import languages
from tqdm import tqdm


def excel2Dict(filename):
    wb = load_workbook(filename=filename, read_only=True)
    ws = wb.active
    records = []
    count = 0
    keys = []
    for row in ws.rows:
        temp = []
        for cell in row:
            if count == 0:
                keys.append(str(cell.value))
            else:
                temp.append(str(cell.value))
        if count > 0:
            records.append(dict(zip(keys, temp)))
        else:
            count += 1
    return records


def saveExcel(filename, title, data):
    wb = Workbook()
    ws = wb.active
    ws.title = title

    count = 0
    for item in data:
        if count == 0:
            ws.append(item.keys())
            count += 1
        ws.append(item.values())
    date = datetime.now().strftime('%Y-%m-%d')
    name = "%s-%s.xlsx" % (filename, date)
    wb.save(name)


def saveWorksheet(wb, title, data, searchTerm=None, orderedEntries=False, autoFilter=False):
    if orderedEntries is True:
        orderedEntries = ['Full Author Names', 'Authors', 'Publication Date', 'Title', 'Publication Type', 'Journal Title', 'Source', 'Language', 'scopusID', 'PMID', 'Citations', 'Citations in Past Year', 'Citations Rate', 'Country of Origin']
    ws = wb.create_sheet()
    ws.title = title
    count = 0
    for record in data:
        if orderedEntries:
            record = OrderedDict((k, record[k]) for k in orderedEntries)
        if count == 0:
            if searchTerm is not None:
                ws.append([searchTerm])
            ws.append(record.keys())
            count += 1
        output = list()
        for val in record.values():
            if isinstance(val, list):
                # dump = json.dumps(val)
                dump = ", ".join(val)
                output.append(dump)
            else:
                output.append(val)
        ws.append(output)
    if autoFilter and orderedEntries:
        ws.auto_filter.ref = "A2:N%d" % (len(data) + 1)
        for row in range(3, ws.max_row):
            cell = ws.cell(row=row, column=13)
            cell.number_format = "0.0"
    elif autoFilter:
        ws.auto_filter.ref = "A2:D%d" % (len(data) + 1)


def distinctSet(records, title):
    data = Set([])
    for record in records:
        if isinstance(record[title], list):
            for item in record[title]:
                data.add(item)
        else:
            data.add(record[title])
    return data


def languageParse(record):
    if len(record['LA']) > 1:
        return [languages.get(part2b=l.lower()).name for l in record['LA']]
    else:
        langCode = record['LA'][0].lower()
        return languages.get(part2b=langCode).name


def outputYearlyData(records):
    output = dict()
    for record in records:
        year = record['Publication Date'].year
        if year in output.keys():
            output[year] += 1
        else:
            output[year] = 1
    return output


def outputYearlyCitationData(records):
    output = dict()
    for record in records:
        year = record['Publication Date'].year
        if year in output.keys():
            output[year] += record['Citations']
        else:
            output[year] = record['Citations']
    return output
PToutput = ['ENGLISH ABSTRACT', 'Meta-Analysis', 'Controlled Clinical Trial', 'LETTER', 'Clinical Trial, Phase III', 'Review', "Research Support, U.S. Gov't, P.H.S.", 'Guideline', 'Interview', "Research Support, Non-U.S. Gov't", 'Consensus Development Conference', 'Lectures', "Research Support, U.S. Gov't, Non-P.H.S.", 'Published Erratum', 'Clinical Study', 'Overall', 'REVIEW', 'Letter', 'Observational Study', 'Comparative Study', 'Clinical Trial, Phase I', 'Comment', 'Multicenter Study', 'Validation Studies', 'Journal Article', 'JOURNAL ARTICLE', 'Historical Article', 'Evaluation Studies', 'Research Support, N.I.H., Extramural', 'Editorial', 'Retracted Publication', 'CASE REPORTS', 'Classical Article', 'Technical Report', 'Randomized Controlled Trial', 'Video-Audio Media', 'English Abstract', 'News', 'Portraits', 'Clinical Trial, Phase II', 'Research Support, N.I.H., Intramural', 'Introductory Journal Article', 'Congresses', 'Case Reports', 'Practice Guideline', 'Clinical Trial', 'Biography']
PTwant = ["Journal Article", "Letter", "Editorial", "Review", "Meta-Analysis", "Research Support, Non-U.S. Gov't", "Research Support, U.S. Gov't, P.H.S.", "Research Support, U.S. Gov't, Non-P.H.S.", "Research Support, N.I.H., Extramural", "Research Support, N.I.H., Intramural", "Clinical Study", "Observational Study", "Comparative Study", "Multicenter Study", "Validation Studies", "Controlled Clinical Trial", "Clinical Trial", "Clinical Trial, Phase I", "Clinical Trial, Phase II", "Clinical Trial, Phase III", "Randomized Controlled Trial", "Guideline", "Practice Guideline", "Lectures", "Case Reports", "Historical Article", "Evaluation Studies"]
PTdel = ['ENGLISH ABSTRACT', 'LETTER', 'Interview', 'Published Erratum', 'Overall', 'REVIEW', 'Comment', 'JOURNAL ARTICLE', 'CASE REPORTS', 'Classical Article', 'Technical Report', 'Video-Audio Media', 'English Abstract', 'News', 'Portraits', 'Introductory Journal Article']


def outputPubDataPubmed(records):
    output = dict()
    renamer = {'LETTER': 'Letter', 'REVIEW': 'Review', 'JOURNAL ARTICLE': 'Journal Article', 'CASE REPORTS': 'Case Reports'}
    for record in records:
        pubTypes = record['Publication Subset']
        pubTypesRenamed = [renamer.get(pub, pub) for pub in pubTypes]
        for pubType in pubTypesRenamed:
            if pubType in output.keys():
                output[pubType] += 1
            else:
                output[pubType] = 1
    output = {p: output.get(p, 0) for p in PTwant}
    return OrderedDict((k, output[k]) for k in PTwant)

PTscopus = ["Article", "Letter", "Editorial", "Review", "Conference Paper", "Chapter", "Note", "Short Survey"]


def outputPubDataScopus(records):
    output = OrderedDict((k, 0) for k in PTscopus)
    for r in records:
        try:
            pubType = r['Publication Type']
            output[pubType] += 1
        except:
            pass
    return output


def getPubmedIds(term, start=0):
    handle = Entrez.esearch(db="pubmed", retstart=start, retmax=20000, term=term)
    record = Entrez.read(handle)
    return record["IdList"]


def parsePubmed(records, total, array, pmids):
    for r in tqdm(records):
        if 'TI' in r:
            pub = dict()
            pub['Full Author Names'] = r.get('FAU', 'Unknown')
            pub['Authors'] = r.get('AU', 'Unknown')
            try:
                pub['Publication Date'] = datetime.strptime(r['EDAT'][:-6], "%Y/%M/%d").date()
            except ValueError:
                try:
                    pub['Publication Date'] = datetime.strptime(r['EDAT'], "%Y/%M/%d").date()
                except ValueError:
                    print("Date parsing error: %s") % r['EDAT']
                    pub['Publication Date'] = r['EDAT']
            pub['Title'] = r['TI']
            pub['Publication Type'] = r['PT']
            pub['Journal Title'] = r.get('JT', 'Unknown')
            pub['Source'] = r.get('SO', 'Unknown')
            pub['Language'] = languageParse(r)
            pub['PMID'] = r['PMID']
            pub['Citations'] = 0
            pub['Citations in Past Year'] = 0
            pub['Citations Rate'] = 0
            array.append(pub)
            pmids.append(r['PMID'])


def cyberVgamma(records, searchTerm):
    cyberPubs = list()
    gammaPubs = list()
    linacPubs = list()
    novalisPubs = list()
    tomoPubs = list()
    protonPubs = list()
    cyberIds = getPubmedIds("%s AND (Cyberknife OR cyber knife)" % searchTerm)
    gammaIds = getPubmedIds("%s AND (gamma knife OR gammaknife)" % searchTerm)
    linacIds = getPubmedIds("%s AND linac" % searchTerm)
    novalisIds = getPubmedIds("%s AND Novalis" % searchTerm)
    tomoIds = getPubmedIds("%s AND TomoTherapy" % searchTerm)
    protonIds = getPubmedIds("%s AND proton beam" % searchTerm)
    for record in records:
        if record['PMID'] in cyberIds:
            cyberPubs.append(record)
        if record['PMID'] in gammaIds:
            gammaPubs.append(record)
        if record['PMID'] in linacIds:
            linacPubs.append(record)
        if record['PMID'] in novalisIds:
            novalisPubs.append(record)
        if record['PMID'] in tomoIds:
            tomoPubs.append(record)
        if record['PMID'] in protonIds:
            protonPubs.append(record)
    return {
        'cyber': {"records": sorted(cyberPubs, key=itemgetter('Citations Rate', 'Citations'), reverse=True), "yearlyData": outputYearlyData(cyberPubs), "pubTypesData": outputPubDataPubmed(cyberPubs), "pubSubsetsData": outputPubDataScopus(cyberPubs), "query": "%s AND (Cyberknife OR cyber knife)" % searchTerm},
        'gamma': {"records": sorted(gammaPubs, key=itemgetter('Citations Rate', 'Citations'), reverse=True), "yearlyData": outputYearlyData(gammaPubs), "pubTypesData": outputPubDataPubmed(gammaPubs), "pubSubsetsData": outputPubDataScopus(gammaPubs), "query": "%s AND (gamma knife OR gammaknife)" % searchTerm},
        'linac': {"records": sorted(linacPubs, key=itemgetter('Citations Rate', 'Citations'), reverse=True), "yearlyData": outputYearlyData(linacPubs), "pubTypesData": outputPubDataPubmed(linacPubs), "pubSubsetsData": outputPubDataScopus(linacPubs), "query": "%s AND linac" % searchTerm},
        'novalis': {"records": sorted(novalisPubs, key=itemgetter('Citations Rate', 'Citations'), reverse=True), "yearlyData": outputYearlyData(novalisPubs), "pubTypesData": outputPubDataPubmed(novalisPubs), "pubSubsetsData": outputPubDataScopus(novalisPubs), "query": "%s AND Novalis" % searchTerm},
        'tomo': {"records": sorted(tomoPubs, key=itemgetter('Citations Rate', 'Citations'), reverse=True), "yearlyData": outputYearlyData(tomoPubs), "pubTypesData": outputPubDataPubmed(tomoPubs), "pubSubsetsData": outputPubDataScopus(tomoPubs), "query": "%s AND TomoTherapy" % searchTerm},
        'proton': {"records": sorted(protonPubs, key=itemgetter('Citations Rate', 'Citations'), reverse=True), "yearlyData": outputYearlyData(protonPubs), "pubTypesData": outputPubDataPubmed(protonPubs), "pubSubsetsData": outputPubDataScopus(protonPubs), "query": "%s AND Proton beam" % searchTerm}
    }


def pubmedData(wb, records, searchTerm, suffix):
    yearlyData = outputYearlyData(records)
    pubTypesData = outputPubDataPubmed(records)
    pubSubsetsData = outputPubDataScopus(records)

    print("Grabing Subsections")
    subsections = cyberVgamma(records, searchTerm)
    subsections['all'] = {
        'query': searchTerm,
        'records': records,
        'yearlyData': yearlyData,
        'pubSubsetsData': pubSubsetsData,
        'pubTypesData': pubTypesData
    }

    ws = wb.active
    ws.title = 'Overview'
    ws.append([
        'Search Term:',
        searchTerm,
        subsections['gamma']['query'],
        subsections['cyber']['query'],
        subsections['linac']['query'],
        subsections['novalis']['query'],
        subsections['tomo']['query'],
        subsections['proton']['query']
    ])
    ws.append([
        'Total',
        sum(yearlyData.values()),
        sum(subsections['gamma']['yearlyData'].values()),
        sum(subsections['cyber']['yearlyData'].values()),
        sum(subsections['linac']['yearlyData'].values()),
        sum(subsections['novalis']['yearlyData'].values()),
        sum(subsections['tomo']['yearlyData'].values()),
        sum(subsections['proton']['yearlyData'].values())
    ])
    ws.append(['Publication Scopus Subsets'])
    rowNum = 3
    for subset, v in pubSubsetsData.items():
        ws.append([
            subset,
            v,
            subsections['gamma']['pubSubsetsData'].get(subset, 0),
            subsections['cyber']['pubSubsetsData'].get(subset, 0),
            subsections['linac']['pubSubsetsData'].get(subset, 0),
            subsections['novalis']['pubSubsetsData'].get(subset, 0),
            subsections['tomo']['pubSubsetsData'].get(subset, 0),
            subsections['proton']['pubSubsetsData'].get(subset, 0)
        ])
        rowNum += 1
    ws.append(['Publication Pubmed Subsets'])
    rowNum += 1
    for pubType, v in pubTypesData.items():
        ws.append([
            pubType,
            v,
            subsections['gamma']['pubTypesData'].get(pubType, 0),
            subsections['cyber']['pubTypesData'].get(pubType, 0),
            subsections['linac']['pubTypesData'].get(pubType, 0),
            subsections['novalis']['pubTypesData'].get(pubType, 0),
            subsections['tomo']['pubTypesData'].get(pubType, 0),
            subsections['proton']['pubTypesData'].get(pubType, 0)
        ])
        rowNum += 1
    ws.append([
        'Year',
        searchTerm,
        subsections['gamma']['query'],
        subsections['cyber']['query'],
        subsections['linac']['query'],
        subsections['novalis']['query'],
        subsections['tomo']['query'],
        subsections['proton']['query']
    ])
    rowNum += 1
    chartStart = rowNum + 1
    oldestYear = min(yearlyData.keys())
    for year in range(oldestYear, 2017):
        ws.append([
            year,
            yearlyData.get(year, 0),
            subsections['gamma']['yearlyData'].get(year, 0),
            subsections['cyber']['yearlyData'].get(year, 0),
            subsections['linac']['yearlyData'].get(year, 0),
            subsections['novalis']['yearlyData'].get(year, 0),
            subsections['tomo']['yearlyData'].get(year, 0),
            subsections['proton']['yearlyData'].get(year, 0)
        ])
        rowNum += 1

    c1 = LineChart()
    c1.width = 30
    c1.height = 15
    dates = Reference(ws, min_row=chartStart, min_col=1, max_col=1, max_row=rowNum)
    s1 = Reference(ws, min_row=chartStart - 1, min_col=2, max_col=2, max_row=rowNum)
    s2 = Reference(ws, min_row=chartStart - 1, min_col=3, max_col=3, max_row=rowNum)
    s3 = Reference(ws, min_row=chartStart - 1, min_col=4, max_col=4, max_row=rowNum)
    c1.series.append(Series(s1, title_from_data=True))
    c1.series.append(Series(s2, title_from_data=True))
    c1.series.append(Series(s3, title_from_data=True))
    c1.set_categories(dates)
    c1.title = "Articles per year for search: %s" % searchTerm
    c1.y_axis.title = 'Number of Articles'
    c1.x_axis.title = 'Year'
    ws.add_chart(c1, "A%d" % (rowNum + 5))

    saveWorksheet(wb, 'Pubmed Stats', records, searchTerm, orderedEntries=True, autoFilter=True)
    return subsections
