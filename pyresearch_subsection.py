from optparse import OptionParser
from Bio import Entrez
from openpyxl import load_workbook, Workbook
from authors import outputAuthors
from journals import outputJournals
from pubmedHelpers import getPubmedIds, saveWorksheet, pubmedData
import numpy
from operator import itemgetter
from datetime import datetime
from tqdm import tqdm
import warnings
warnings.filterwarnings("ignore")
with open('config.txt', 'r') as f:
    Entrez.email = f.readline()

parser = OptionParser()
parser.add_option("-s", "--search", dest="searchTerm", help="search term")
parser.add_option("-i", "--input", dest="input", help="Name of input file")
parser.add_option("-d", "--data", dest="data", help="Name of numpy file")
parser.add_option("-t", "--title", dest="title", help="Name of worksheet")
(options, args) = parser.parse_args()
if len(options.searchTerm) == 0:
	print("Please input a search term with -s or --search")
	exit(1)
print("Search Term: %s") % options.searchTerm

date = datetime.now().strftime('%Y-%m-%d')
suffix = "-%s" % date

idlist = getPubmedIds(options.searchTerm)
total = len(idlist)
print("Number of articles: %s") % total

data = numpy.load('%s.npy' % options.input)
matched = list()
for record in tqdm(data):
	if record['PMID'] in idlist:
		matched.append(record)
publications = sorted(matched, key=itemgetter('Citations Rate', 'Citations'), reverse=True)

print('Saving to excel')
wb = load_workbook('%s.xlsx' % options.input)
saveWorksheet(wb, options.title, publications, options.searchTerm, orderedEntries=True, autoFilter=True)

wb.save('%s.xlsx' % options.input)
directory = 'subsections'
numpy.save('%s/%s.npy' % (directory, options.title), publications)

# Subjection Excel file output

searchTerms = {
    "Brain Metastases": "Stereotactic AND (radiosurgery OR radiotherapy) AND (brain OR cranial OR cranium) AND (metastasis OR metastases)",
    "Spinal Metastases": "Stereotactic AND (radiosurgery OR radiotherapy) AND (spine OR brainstem OR spinal cord) AND (metastasis OR metastases)",
    "Meningioma": "Stereotactic AND (radiosurgery OR radiotherapy) AND Meningioma",
    "Glioblastoma": "Stereotactic AND (radiosurgery OR radiotherapy) AND (Glioblastoma OR GBM OR high grade glioma OR astrocytoma oligodendroglioma)",
    "AVM": "Stereotactic AND (radiosurgery OR radiotherapy) AND (Arteriovenous malformation OR AVM)",
    "Acoustic Neuroma": "Stereotactic AND (radiosurgery OR radiotherapy) AND (Acoustic neuroma OR vestibular schwanoma)",
    "Trigeminal Neuralgia": "Stereotactic AND (radiosurgery OR radiotherapy) AND (Trigeminal Neuralgia OR Tic Douloureux)",
    "Psychiatry": "Stereotactic AND (radiosurgery OR radiotherapy) AND (depression OR anxiety disorder OR obsessive compulsive OR obsessive compulsion)"
}
searchTerm = searchTerms[options.title]

wb = Workbook()
subsections = pubmedData(wb, publications, searchTerm, suffix)
numpy.save('%s/%s-subsections.npy' % (directory, options.title), subsections)

authorData = outputAuthors(publications)
saveWorksheet(wb, 'Authors', authorData, searchTerm=searchTerm, autoFilter=True)

journalData = outputJournals(publications)
saveWorksheet(wb, 'Journals', journalData)

saveWorksheet(wb, 'Gamma Knife', subsections['gamma']['records'], subsections['gamma']['query'], orderedEntries=True, autoFilter=True)
saveWorksheet(wb, 'CyberKnife', subsections['cyber']['records'], subsections['cyber']['query'], orderedEntries=True, autoFilter=True)
saveWorksheet(wb, 'linac', subsections['linac']['records'], subsections['linac']['query'], orderedEntries=True, autoFilter=True)
saveWorksheet(wb, 'Novalis', subsections['novalis']['records'], subsections['novalis']['query'], orderedEntries=True, autoFilter=True)
saveWorksheet(wb, 'TomoTherapy', subsections['tomo']['records'], subsections['tomo']['query'], orderedEntries=True, autoFilter=True)
wb.save('%s/%s.xlsx' % (directory, options.title))
